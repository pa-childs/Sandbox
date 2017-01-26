#! python3
"""
Project:    Sandbox
Filename:   openpyxl_test
Created by: PJC
Created on: July 13, 2016
"""

import datetime
import logging
import os
import openpyxl
# from openpyxl.cell import column_index_from_string
# from openpyxl.cell import get_column_letter

# Setup logging.  Display INFO messages and higher to console
logger = logging.getLogger(__name__)
log_handler = logging.StreamHandler()
logger.setLevel(logging.INFO)
log_handler.setLevel(logging.INFO)
logger.addHandler(log_handler)

test_directory = 'C:\\Work_Files\\PyCharm\\Sandbox\\TestFiles\\example_OpenPyXL\\'
test_file = 'Test_Workbook.xlsx'


def main():

    # Open test workbook

    try:

        os.chdir(test_directory)
        workbook_file = openpyxl.load_workbook(test_file)

    except IOError as e:

        logger.error('Unable to open file: {0}'.format(e))
        quit()

    # List all sheet names
    sheet_names = workbook_file.get_sheet_names()
    logger.info('List of sheetnames: {0}'.format(sheet_names))

    # Assign all sheets to objects
    sheet_one = workbook_file.get_sheet_by_name('Test Sheet')
    logger.info('Retrieved sheet_one: {0}'.format(sheet_one.title))

    sheet_two = workbook_file.get_sheet_by_name('Sheet2')
    logger.info('Retrieved sheet_two: {0}'.format(sheet_two.title))

    sheet_three = workbook_file.get_sheet_by_name('Sheet3')
    logger.info('Retrieved sheet_two: {0}'.format(sheet_three.title))

    # Open a sheet that doesn't exist
    try:

        sheet_four = workbook_file.get_sheet_by_name('Sheet4')
        logger.info('Retrieved sheet_four: {0}'.format(sheet_four.title))

    except KeyError as e:

        logger.info(e)

    # Show which sheet is active
    logger.info('Active sheet is: {0}\n'.format(workbook_file.active))

    # Show values for a Cell in sheet_one using coordinates
    cell_obj = sheet_one['B1']
    logger.info('Cell value is: {0}'.format(cell_obj.value))
    logger.info('The Cell is in column {0}'.format(cell_obj.column))
    logger.info('The Cell is in row {0}'.format(cell_obj.row))
    logger.info('The Cell coordinate is {0}\n'.format(cell_obj.coordinate))

    # Show values for a Cell in sheet_one using row and column numbers
    cell_obj = sheet_one.cell(row=7, column=2).value
    logger.info('Cell value is: {0}'.format(cell_obj))

    for i in range(1, 8, 2):

        if i == 7:

            logger.info('Row {0}: {1}\n'.format(i, sheet_one.cell(row=i, column=2).value))

        else:

            logger.info('Row {0}: {1}'.format(i, sheet_one.cell(row=i, column=2).value))

    # Convert column letters and numbers
    column_number = column_index_from_string('BZ')
    logger.info('Column "BZ" converts to {0}'.format(column_number))

    column_letter = get_column_letter(27)
    logger.info('Column "27" converts to {0}\n'.format(column_letter))

    # Access all Cells in a column
    column_b = sheet_one.columns[1]
    for cell_obj in column_b:
        logger.info(cell_obj.value)  # Extra column exists because openpyxl can't delete empty edited cells
    logger.info('\n')

    try:

        # Update spreadsheet with new values
        # Update title for a sheet
        sheet_three.title = 'New Title'
        workbook_file.save(test_file)
        logger.info('Changed sheet_three title to: {0}'.format(sheet_three.title))

        # Create a new sheet
        workbook_file.create_sheet('Sheet4')
        sheet_four = workbook_file.get_sheet_by_name('Sheet4')
        workbook_file.save(test_file)
        logger.info('Created sheet_four: {0}'.format(sheet_four.title))

        # Add values to a sheet
        sheet_one['A8'] = datetime.datetime.today()
        sheet_one['B8'] = 'Raspberries'
        sheet_one['C8'] = 42
        workbook_file.save(test_file)
        logger.info('Added Cells contain the values: {0}, {1}, {2}'.format(sheet_one['A8'].value,
                                                                           sheet_one['B8'].value,
                                                                           sheet_one['C8'].value))

        # Reset Values
        sheet_three.title = 'Sheet3'
        workbook_file.remove_sheet(sheet_four)
        # You can't really delete the cell values
        sheet_one['A8'] = ''
        sheet_one['B8'] = ''
        sheet_one['C8'] = ''

        workbook_file.save(test_file)
        logger.info('Reset {0} to initial values.'.format(test_file))

    except PermissionError as e:

        logger.error('Unable to update spreadsheet: {0}'.format(e))

if __name__ == '__main__':
    main()
